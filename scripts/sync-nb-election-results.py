#!/usr/bin/env python3
"""
New Brunswick Election Results Sync — Local Government Mayors & Councillors

Source of truth for NB: Elections NB's own official May 2026 election
publications (electionsnb.ca > Resources > Results/Reports/Publications >
Municipal, District Education Council, Regional Health Authority Elections):
  - "Lists of Candidates" PDF -- every nominated candidate per race, with
    acclaimed winners explicitly marked "(accl.)"
  - "Results (Tab)" XLSX -- per-poll tabulator vote counts for every
    contested race, with a "Total" row per candidate

Why this instead of OpenNorth: NB held a general local election on May 11,
2026 -- our existing ~592 "current" NB office_holders rows (from OpenNorth)
predate it and are almost certainly showing the outgoing council, the same
bug class as Maple Ridge but province-wide and far more recent. NB also
underwent a major local governance reform in 2023 that collapsed hundreds
of small entities into ~81 new "local governments" (the ones this script
covers) plus a separate, larger set of new "rural districts" that elect
advisory committees, not a mayor/council -- see the note on rural districts
below.

Winner determination:
  - A candidate marked "(accl.)" in the PDF won by acclamation (only
    nominee) -- no vote count exists for these races at all, since nothing
    was tabulated.
  - Otherwise, the winner(s) are the top N vote-getters (N = seat count,
    from the XLSX "Vote for N" contest header) from that race's XLSX
    "Total" row.

Known gap -- NOT covered by this script: NB's new "rural districts" (a
different, non-municipality category created by the 2023 reform) elect
"Rural District Advisory Committee" members, not mayors/councillors --
that's a different role with no election_role_types row today. Some
fraction of the 271 New Brunswick rows in map_shapes (boundary_type
'Municipal') are likely these rural districts rather than true local
governments; this script only touches shapes whose name matches an actual
Local Government Candidates entry, so rural district rows are left alone
rather than guessed at.

Usage:
  python3 scripts/sync-nb-election-results.py [--apply]

  --apply  Actually run the generated SQL against the database. Without
           it, the script only fetches/parses/prints a summary and writes
           the SQL to scripts/nb-election-results-sync.sql for review.
"""

import argparse
import re
import subprocess
import zipfile
import io
from curl_cffi import requests
import pdfplumber
import openpyxl

DB_URI = "postgresql://postgres:pa.8tX5%2BHh%2FGZn2@db.qlzyfdwrkcxyqapewxwg.supabase.co:5432/postgres"

MAYOR_ROLE_ID = "3855aeb9-c840-4d81-bb83-66cad128ea8c"
COUNCILLOR_ROLE_ID = "6d2e3815-05d9-4784-b220-1d73140b5bf3"

CANDIDATES_PDF_URL = "https://www.electionsnb.ca/content/dam/enb/pdf/event-may-2026/lists-of-candidates.pdf"
RESULTS_ZIP_URL = "https://www.electionsnb.ca/content/dam/enb/xls/mun2026/2026-05-11-MUN-Results-Resultats.zip"
RESULTS_XLSX_NAME = "2026-05-11 Results-Resultats (Tab).xlsx"
SOURCE_URL = "https://electionsnb.ca/content/enb/en/results-reports-publications.html"

# Page-footer boilerplate (repeated on every page, bilingual, wraps/splits
# across the 3-column crop in ways that would otherwise get mistaken for a
# municipality name and corrupt every candidate parsed until the next real
# municipality header shows up).
NOISE_SUBSTRINGS = [
    "electionsnb.ca", "electi", "ionsnb", "acadie nouvelle", "avis en fran",
    "affich", "1-888-858", "v0te", "www.electi", "vote)",
]


def is_noise(line):
    low = line.lower()
    if any(s in low for s in NOISE_SUBSTRINGS):
        return True
    if re.match(r"^[A-Za-z]+ \d{1,2}, \d{4}$", line):  # "May 11, 2026"
        return True
    return False


def _normalize_one(s):
    s = s.lower().strip()
    s = re.sub(r"^the\s+", "", s)
    s = re.sub(
        r"^(city|town|township|municipality|municipalit[eé]|ville|district|borough|county|region|village|"
        r"rm of|resort village of|district of|regional community of|rural community of|"
        r"municipal district of|community of)\s+(of\s+|des\s+|de\s+)?",
        "",
        s,
    )
    s = re.sub(r"\s+(city|town|township|municipality|ville|district|borough|county|region|village)$", "", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return s.strip()


def normalize_municipal_name(s):
    """
    Returns every normalized key a name could plausibly be looked up under.
    New Brunswick's GeoNB boundary names sometimes carry a corporate-form
    prefix ("The City of Fredericton", "Village of Doaktown") that Elections
    NB's plain-English candidate lists don't, and some carry a bilingual
    "English / French" pair ("Grand-Sault / Grand Falls") -- both sides are
    returned as candidate keys so a match succeeds regardless of which
    variant the other source used.
    """
    if not s:
        return []
    parts = [p.strip() for p in s.split("/") if p.strip()]
    return [_normalize_one(p) for p in parts] or [_normalize_one(s)]


def fetch_candidates_pdf():
    print("Fetching Elections NB candidates list...")
    r = requests.get(CANDIDATES_PDF_URL, impersonate="chrome")
    munis = {}
    current_muni = None
    with pdfplumber.open(io.BytesIO(r.content)) as pdf:
        # Pages after the "Local Government Candidates" section (index 1-4
        # in the 2026 PDF) cover Rural District Advisory Committees and
        # District Education Councils -- different categories, deliberately
        # excluded. If Elections NB reshuffles page order in a future year,
        # this cutoff needs re-checking against the section header text.
        for pageno, page in enumerate(pdf.pages):
            if pageno == 0 or pageno > 4:
                continue
            for (x0, x1) in [(45, 265), (265, 485), (485, 744)]:
                crop = page.crop((x0, 100, x1, page.height))
                txt = crop.extract_text() or ""
                for line in txt.split("\n"):
                    line = line.strip()
                    if not line or line.startswith("Position") or is_noise(line):
                        continue
                    m = re.match(
                        r"^(Mayor|Councillor(?: at Large| Ward [\w\-]+(?:\s*-\s*[\w\s\.']+)?)?)\s+(.+?)(\s*\(accl\.\))?$",
                        line,
                    )
                    if m:
                        position, name, accl = m.group(1).strip(), m.group(2).strip(), bool(m.group(3))
                        if current_muni is None:
                            continue
                        munis.setdefault(current_muni, []).append((position, name, accl))
                    else:
                        current_muni = line
    return munis


def fetch_results_xlsx():
    print("Fetching Elections NB per-poll results...")
    r = requests.get(RESULTS_ZIP_URL, impersonate="chrome")
    with zipfile.ZipFile(io.BytesIO(r.content)) as z:
        with z.open(RESULTS_XLSX_NAME) as f:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    contests = {}
    i = 0
    while i < len(rows):
        cell = rows[i][0]
        if isinstance(cell, str) and "Vote for" in cell:
            m = re.match(r"^(.*?),\s*([^,]+?)\s*\(Vote for\s*/?\s*Voter pour\s*(\d+)\)", cell)
            i += 1
            while i < len(rows) and all(c is None for c in rows[i]):
                i += 1
            if i >= len(rows):
                break
            candidates = [c for c in rows[i][1:] if c]
            i += 1
            total_row = None
            while i < len(rows) and rows[i][0] is not None:
                if str(rows[i][0]).strip() == "Total":
                    total_row = rows[i]
                i += 1
            if total_row and m:
                pos_text, muni, seats = m.group(1), m.group(2).strip(), int(m.group(3))
                votes = {cand: total_row[idx + 1] for idx, cand in enumerate(candidates) if total_row[idx + 1] is not None}
                contests[(muni, pos_text.split(" / ")[0].strip())] = (seats, votes)
        else:
            i += 1
    return contests


def resolve_winners(munis, contests):
    winners_by_muni = {}
    unresolved = []
    for muni, cands in munis.items():
        by_pos = {}
        for pos, name, accl in cands:
            by_pos.setdefault(pos, []).append((name, accl))
        for pos, entries in by_pos.items():
            accl_winners = [n for n, a in entries if a]
            if accl_winners:
                winners_by_muni.setdefault(muni, []).extend([(pos, n) for n in accl_winners])
                continue
            key = (muni, pos)
            if key in contests:
                seats, votes = contests[key]
                ranked = sorted(votes.items(), key=lambda kv: -kv[1])
                winners_by_muni.setdefault(muni, []).extend([(pos, n) for n, v in ranked[:seats]])
            else:
                unresolved.append((muni, pos, [n for n, a in entries]))
    return winners_by_muni, unresolved


def run_sql(sql, capture=True):
    cmd = ["psql", DB_URI]
    if capture:
        cmd += ["-t", "-A", "-F", "\t"]
    cmd += ["-c", sql]
    res = subprocess.run(cmd, capture_output=capture, text=True)
    if res.returncode != 0:
        raise RuntimeError(f"psql error: {res.stderr}")
    if capture:
        return [l.split("\t") for l in res.stdout.strip().split("\n") if l.strip()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    munis = fetch_candidates_pdf()
    print(f"Parsed {len(munis)} local governments, {sum(len(v) for v in munis.values())} candidate rows from the PDF.")
    contests = fetch_results_xlsx()
    print(f"Parsed {len(contests)} contested-race vote totals from the XLSX.")

    winners_by_muni, unresolved = resolve_winners(munis, contests)
    total_winners = sum(len(v) for v in winners_by_muni.values())
    print(f"Resolved {total_winners} winners across {len(winners_by_muni)} municipalities.")
    if unresolved:
        print(f"{len(unresolved)} races could not be resolved (name-matching edge cases, e.g. multi-word hyphenated ward names) -- not silently dropped:")
        for muni, pos, names in unresolved:
            print(f"  [{muni}] {pos}: {names}")

    print("Loading map_shapes (Municipal, New Brunswick, active only)...")
    shapes = run_sql("""
        SELECT ms.id, ms.name FROM map_shapes ms
        JOIN shape_containers sc ON sc.map_shape_id = ms.id
        JOIN map_shapes p ON p.id = sc.container_shape_id AND p.name = 'New Brunswick'
        WHERE ms.boundary_type = 'Municipal' AND ms.retired_at IS NULL;
    """)
    shape_by_norm = {}
    dupes = set()
    for sid, name in shapes:
        for norm in normalize_municipal_name(name):
            if norm in shape_by_norm and shape_by_norm[norm] != sid:
                dupes.add(name)
            shape_by_norm[norm] = sid

    unmatched_munis = []
    value_rows = []
    for muni, winners in winners_by_muni.items():
        shape_id = None
        for norm in normalize_municipal_name(muni):
            shape_id = shape_by_norm.get(norm)
            if shape_id:
                break
        if not shape_id:
            unmatched_munis.append(muni)
            continue
        for pos, name in winners:
            role_id = MAYOR_ROLE_ID if pos == "Mayor" else COUNCILLOR_ROLE_ID
            safe_name = name.replace("'", "''")
            value_rows.append(f"({shape_id}, '{role_id}', '{safe_name}', '{SOURCE_URL}')")

    if unmatched_munis:
        print(f"{len(unmatched_munis)} municipalities from the PDF have no matching map_shapes row (likely a coverage gap post-2023 reform, or a name-format mismatch) -- not silently dropped:")
        for m in sorted(unmatched_munis):
            print(f"  {m}")
    if dupes:
        print(f"map_shapes has duplicate names for {len(dupes)} New Brunswick municipalities (data-quality issue, separate from this script) -- an arbitrary one of each was used: {sorted(dupes)}")

    sql = ["BEGIN;"]
    sql.append("""
CREATE TEMP TABLE staging_nb_winners (
  map_shape_id bigint,
  election_role_type_id uuid,
  full_name text,
  source_url text
) ON COMMIT DROP;
""")
    sql.append("INSERT INTO staging_nb_winners VALUES\n" + ",\n".join(value_rows) + ";")

    sql.append("""
UPDATE office_holders oh
SET is_current = false, term_ended_at = CURRENT_DATE, updated_at = NOW()
WHERE oh.is_current = true
  AND EXISTS (SELECT 1 FROM staging_nb_winners s WHERE s.map_shape_id = oh.map_shape_id)
  AND NOT EXISTS (
    -- Matched on (shape, name, role), not just (shape, name) -- see the
    -- identical fix in sync-bc-election-results.py for why: matching by
    -- name alone leaves a stale row behind whenever fresh data confirms
    -- someone in a different role than an old row already had (a role
    -- change retires the old role's row and the new role gets its own
    -- inserted row, same as this would do anyway -- there's no case this
    -- makes worse).
    SELECT 1 FROM staging_nb_winners s
    WHERE s.map_shape_id = oh.map_shape_id AND s.full_name = oh.full_name
      AND s.election_role_type_id = oh.election_role_type_id
  );
""")

    sql.append("""
INSERT INTO office_holders (
  map_shape_id, election_role_type_id, full_name, bio, source_url,
  holding_since, is_current, term_ended_at, updated_at
)
SELECT DISTINCT ON (s.map_shape_id, s.election_role_type_id, s.full_name)
  s.map_shape_id, s.election_role_type_id, s.full_name,
  (SELECT ert.role_title FROM election_role_types ert WHERE ert.id = s.election_role_type_id) || ' for ' ||
  (SELECT ms.name FROM map_shapes ms WHERE ms.id = s.map_shape_id),
  s.source_url, '2026-05-11', true, NULL, NOW()
FROM staging_nb_winners s
ORDER BY s.map_shape_id, s.election_role_type_id, s.full_name
ON CONFLICT (map_shape_id, election_role_type_id, full_name) DO UPDATE SET
  source_url = EXCLUDED.source_url,
  holding_since = EXCLUDED.holding_since,
  is_current = true,
  term_ended_at = NULL,
  updated_at = NOW();
""")

    sql.append("""
DO $$
DECLARE
  r RECORD;
  new_profile_id UUID;
  new_ghost_id UUID;
  existing_profile_id UUID;
  computed_slug TEXT;
  created_count INT := 0;
  linked_count INT := 0;
BEGIN
  FOR r IN
    SELECT oh.id as office_holder_id, oh.full_name, oh.bio,
           ms.country, ms.name as boundary_name, ms.boundary_type, ert.role_title
    FROM office_holders oh
    JOIN map_shapes ms ON oh.map_shape_id = ms.id
    JOIN election_role_types ert ON oh.election_role_type_id = ert.id
    JOIN staging_nb_winners s ON s.map_shape_id = oh.map_shape_id AND s.full_name = oh.full_name
    WHERE oh.linked_profile_id IS NULL
  LOOP
    computed_slug := lower(regexp_replace(regexp_replace(r.full_name || '-' || r.role_title, '[^a-zA-Z0-9]+', '-', 'g'), '(^-|-$)', '', 'g'));
    SELECT pp.id INTO existing_profile_id FROM politician_profiles pp WHERE pp.wall_slug = computed_slug LIMIT 1;
    IF existing_profile_id IS NULL THEN
      SELECT p.id INTO existing_profile_id FROM profiles p
      WHERE p.role = 'politician' AND lower(p.full_name) = lower(r.full_name) AND p.constituency = r.boundary_name
      LIMIT 1;
    END IF;
    IF existing_profile_id IS NOT NULL THEN
      UPDATE office_holders SET linked_profile_id = existing_profile_id WHERE id = r.office_holder_id;
      linked_count := linked_count + 1;
    ELSE
      new_profile_id := gen_random_uuid();
      new_ghost_id := gen_random_uuid();
      INSERT INTO profiles (id, role, full_name, country, constituency, designation, current_ghost_id, updated_at)
      VALUES (new_profile_id, 'politician', r.full_name, r.country, r.boundary_name, r.role_title, new_ghost_id, NOW());
      INSERT INTO politician_profiles (id, political_target_role, target_boundary_type, target_boundary_name, bio, wall_slug, created_at, updated_at)
      VALUES (new_profile_id, r.role_title, r.boundary_type, r.boundary_name, r.bio, computed_slug, NOW(), NOW());
      UPDATE office_holders SET linked_profile_id = new_profile_id WHERE id = r.office_holder_id;
      created_count := created_count + 1;
    END IF;
  END LOOP;
  RAISE NOTICE 'NB sync: created % new ghost profile walls, linked % to existing profiles.', created_count, linked_count;
END $$;
""")
    sql.append("COMMIT;")
    full_sql = "\n".join(sql)

    out_path = "/Users/vmn2k4/Coding/Choseno/scripts/nb-election-results-sync.sql"
    with open(out_path, "w") as f:
        f.write(full_sql)
    print(f"Wrote generated SQL to {out_path}")

    if args.apply:
        print("Applying to database...")
        subprocess.run(["psql", DB_URI, "-f", out_path], check=True)
        print("Done.")
    else:
        print("Dry run only (no --apply passed) -- review the SQL file, then re-run with --apply.")


if __name__ == "__main__":
    main()
